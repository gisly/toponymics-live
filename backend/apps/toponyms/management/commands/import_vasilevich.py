"""
Импортёр топонимических данных Г.М. Василевич из xlsx.

Использование:
    docker compose exec django python manage.py import_vasilevich \\
        --file /data-migration/Vasilevich_data_for_platform.xlsx \\
        [--dry-run] [--update-existing]

Что делает:
1. Открывает Excel — 17 вкладок, каждая = одна рукописная карта
2. Создаёт справочники (Language, FeatureType, Person, SourceReference)
3. Создаёт HistoricalMap для каждой вкладки + Г.М. Василевич как collector
4. Для каждой строки с непустым именем создаёт Place + Toponym
5. Если есть Official names — создаёт второй Toponym на русском с тем же Place
6. Пишет markdown-лог с тем что куда легло и warnings

Idempotency: повторный запуск с --update-existing обновит существующие
записи; без флага пропустит уже импортированные строки.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from apps.toponyms.models import (
    FeatureType,
    HistoricalMap,
    Language,
    Person,
    Place,
    SourceReference,
    Toponym,
)


# ─── Константы ─────────────────────────────────────────────────────────


COLLECTOR_NAME = "Василевич Г.М."

ARCHIVAL_SOURCE = (
    "Научный архив МАЭ РАН. Ф.22. Оп.2 Д.74-75. "
    "Василевич Г.М. Карты рек, выполненные от руки тушью и карандашом. "
    "Бумага, калька. [1924 – 1969]"
)

# Маппинг "сырого" значения языка → ISO код
LANGUAGE_NORMALIZATION = {
    "Evenki": "evn",
    "Evenk": "evn",         # опечатка — приводим к evn
    "эвенкийский": "evn",
    "Yakut": "sah",
    "Yakut ?": "sah",       # с вопросом — всё равно sah
    "Russian": "ru",
    "Russian (?)": "ru",
    "Keto": "ket",          # кетский
}

LANGUAGE_FULL_DATA = {
    "evn": {"name_ru": "Эвенкийский", "name_en": "Evenki", "name_native": "Эвэды̄ турэ̄н"},
    "sah": {"name_ru": "Якутский", "name_en": "Yakut", "name_native": "Саха тыла"},
    "ru":  {"name_ru": "Русский", "name_en": "Russian", "name_native": "Русский"},
    "ket": {"name_ru": "Кетский", "name_en": "Ket", "name_native": "Кэтскай"},
    "en":  {"name_ru": "Английский", "name_en": "English", "name_native": "English"},
}


# Нормализация источника координат (опечатки в данных)
COORD_SOURCE_NORMALIZATION = {
    "GGC500": "GGC500m",
    "GGC501": "GGC500m",
}


# Имена колонок, которые мы ищем в заголовке (строка 3)
COLUMN_HEADERS = {
    "name_evn": ["place names in Evenki"],
    "name_lat": ["transliteration"],
    "translation_ru": ["translation in Russian"],
    "translation_en": ["translation in English"],
    "affix": ["Affix"],
    "obj_ru": ["Объект"],
    "obj_en": ["Object"],
    "obj_evn_cyr": ["Объект (эвенкийский)"],
    "obj_evn_lat": ["Object (Evenki)"],
    "language": ["Language"],
    "map_yesno": ["Map (yes / no)"],
    "comment": ["Comment"],
    "official_name": ["Official names"],
    "latitude": ["Latitude"],
    "longitude": ["Longitude"],
    "precise": ["Precise"],
    "source": ["Source"],
    "author": ["author"],
}


@dataclass
class ImportStats:
    sheets_processed: int = 0
    maps_created: int = 0
    maps_updated: int = 0
    maps_skipped: int = 0
    places_created: int = 0
    toponyms_evn_created: int = 0
    toponyms_ru_created: int = 0
    rows_skipped_empty: int = 0
    rows_no_coords: int = 0
    warnings: list[str] = field(default_factory=list)


class Command(BaseCommand):
    help = "Импорт топонимических данных Василевич из Excel"

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, type=Path)
        parser.add_argument("--dry-run", action="store_true",
                            help="Не сохранять в БД, только показать что будет")
        parser.add_argument("--update-existing", action="store_true",
                            help="Обновить существующие записи")
        parser.add_argument("--log-output", type=Path, default=None)

    def handle(self, *args, **options):
        self.file_path: Path = options["file"]
        self.dry_run: bool = options["dry_run"]
        self.update_existing: bool = options["update_existing"]
        self.log_output = options["log_output"] or (
            self.file_path.parent / f"vasilevich_import_log_{datetime.now():%Y%m%d_%H%M%S}.md"
        )

        if not self.file_path.exists():
            raise CommandError(f"Файл не найден: {self.file_path}")

        self.stats = ImportStats()
        self.log_lines: list[str] = []

        self._log_header(f"Импорт Vasilevich data → новая БД")
        self.log(f"- Файл: `{self.file_path.absolute()}`")
        self.log(f"- Dry-run: **{'да' if self.dry_run else 'нет'}**")
        self.log(f"- Обновлять существующие: {'да' if self.update_existing else 'нет'}")

        # Прежде всего готовим справочники: языки, общий источник, collector
        self._setup_reference_data()

        # Открываем книгу
        wb = load_workbook(self.file_path, data_only=True)
        self.log(f"\nЗагружено {len(wb.sheetnames)} вкладок (= рукописных карт)")

        # Обрабатываем каждую вкладку
        for sheet_name in wb.sheetnames:
            try:
                self._import_sheet(wb[sheet_name], sheet_name)
            except Exception as e:
                self.stats.warnings.append(f"Лист '{sheet_name}': {e}")
                self.log(f"\n  ❌ **Ошибка на листе '{sheet_name}'**: {e}")

        # Финальная сводка
        self._write_summary()
        self._save_log()

        self.stdout.write(self.style.SUCCESS(
            f"\n✓ Импорт завершён. Лог: {self.log_output}"
        ))

    # ─── Справочники ──────────────────────────────────────────────────

    def _setup_reference_data(self):
        """Создаёт языки, общий источник, collector (Василевич Г.М.)."""
        self._log_header("Шаг 1: справочники")

        # Языки
        self.languages: dict[str, Language] = {}
        for iso, data in LANGUAGE_FULL_DATA.items():
            if self.dry_run:
                self.languages[iso] = Language(iso=iso, **data)
                continue
            lang, created = Language.objects.update_or_create(iso=iso, defaults=data)
            self.languages[iso] = lang
            self.log(f"  {'+ создан' if created else '= уже был'} язык {iso} ({data['name_ru']})")

        # Общий источник архивных данных Василевич
        if self.dry_run:
            self.archival_source = SourceReference(description=ARCHIVAL_SOURCE)
        else:
            self.archival_source, _ = SourceReference.objects.get_or_create(
                description=ARCHIVAL_SOURCE,
            )
            self.log(f"  + источник: «{ARCHIVAL_SOURCE[:60]}…»")

        # Г.М. Василевич как collector
        if self.dry_run:
            self.collector = Person(
                last_name="Василевич", first_name="Глафира", patronymic="Макарьевна",
                comment="Этнограф, лингвист, исследователь эвенкийского языка и культуры",
            )
        else:
            self.collector, _ = Person.objects.get_or_create(
                last_name="Василевич", first_name="Глафира", patronymic="Макарьевна",
                defaults={"comment": "Этнограф, лингвист, исследователь эвенкийского языка и культуры"},
            )
            self.log(f"  + collector: {self.collector.full_name}")

        # FeatureType кэш (создаём ленивым способом по мере встречи)
        self.feature_types: dict[str, FeatureType] = {}
        if not self.dry_run:
            for ft in FeatureType.objects.all():
                self.feature_types[ft.code] = ft

    def _get_or_create_feature_type(self, ru: str | None, en: str | None,
                                    evn_cyr: str | None = None,
                                    evn_lat: str | None = None) -> FeatureType | None:
        """Находит или создаёт FeatureType по русскому+английскому названию.

        Если есть эвенкийское имя — сохраняет его в description_ru как примечание.
        """
        if not ru and not en:
            return None

        # Нормализуем длинные английские варианты "yar (an outcrop ...)" в просто "yar"
        if en:
            en = en.split("(")[0].strip() or en

        code = self._slugify_simple(en or ru)
        if not code:
            return None

        if code in self.feature_types:
            return self.feature_types[code]

        if self.dry_run:
            ft = FeatureType(
                code=code, name_ru=ru or en, name_en=en or ru,
                description_ru=f"Эвенкийское: {evn_cyr or ''}" if evn_cyr else "",
            )
            self.feature_types[code] = ft
            return ft

        ft, created = FeatureType.objects.get_or_create(
            code=code,
            defaults={
                "name_ru": ru or en,
                "name_en": en or ru,
                "description_ru": f"Эвенкийское название: {evn_cyr}" if evn_cyr else "",
                "description_en": f"Evenki: {evn_lat}" if evn_lat else "",
                "language": self.languages["evn"] if evn_cyr else None,
            },
        )
        self.feature_types[code] = ft
        if created:
            self.log(f"  + тип объекта: {code} ({ru} / {en})")
        return ft

    @staticmethod
    def _slugify_simple(text: str) -> str:
        """Преобразует строку в slug-вид (только ASCII буквы и цифры)."""
        text = (text or "").lower().strip()
        # Берём только латинские буквы/цифры, остальное → дефис
        slug = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
        return slug[:50]

    # ─── Обработка одной вкладки ─────────────────────────────────────

    def _import_sheet(self, sheet, sheet_name: str):
        """Импортирует одну вкладку — одну рукописную карту."""
        self._log_header(f"Лист: '{sheet_name}'")
        self.stats.sheets_processed += 1

        # Парсим имя карты из имени вкладки: "map 14; 22-2-75-2"
        m = re.match(r"^\s*map\s+(\S+?);\s*(.+?)\s*$", sheet_name, re.IGNORECASE)
        if m:
            map_number = m.group(1)
            archive_ref = m.group(2)
        else:
            map_number = sheet_name
            archive_ref = sheet_name

        # Заголовки колонок (строка 3)
        header_map = self._find_headers(sheet)
        if not header_map.get("name_evn"):
            self.log(f"  ⚠ Не нашли колонку 'place names in Evenki' — пропуск")
            return

        # Информант — строка 4, колонка author
        informant_raw = self._cell(sheet, 4, header_map.get("author"))
        informant = self._get_or_create_person(informant_raw)
        if informant:
            self.log(f"  Информант: {informant.full_name}")

        # Создаём (или обновляем) HistoricalMap
        historical_map = self._get_or_create_map(map_number, archive_ref, informant)
        if not historical_map:
            return

        # Идём по строкам с данными — начинаем со строки 4 (там могут быть данные)
        rows_in_sheet = 0
        for row_idx in range(4, sheet.max_row + 1):
            name_evn = self._cell(sheet, row_idx, header_map.get("name_evn"))
            if not name_evn:
                self.stats.rows_skipped_empty += 1
                continue

            try:
                self._import_row(
                    sheet, row_idx, header_map,
                    historical_map=historical_map,
                    informant=informant,
                )
                rows_in_sheet += 1
            except Exception as e:
                self.stats.warnings.append(
                    f"Лист '{sheet_name}' строка {row_idx}: {e}"
                )

        self.log(f"  Импортировано строк: **{rows_in_sheet}**")

    def _find_headers(self, sheet) -> dict[str, int | None]:
        """Возвращает {смысловое_имя: col_idx} для каждого известного заголовка."""
        result = {key: None for key in COLUMN_HEADERS}
        for col_idx in range(1, sheet.max_column + 1):
            v = sheet.cell(row=3, column=col_idx).value
            if not v:
                continue
            v_str = str(v).strip()
            for key, candidates in COLUMN_HEADERS.items():
                if v_str in candidates:
                    result[key] = col_idx
                    break
        return result

    def _cell(self, sheet, row: int, col: int | None):
        if not col:
            return None
        v = sheet.cell(row=row, column=col).value
        if v is None:
            return None
        if isinstance(v, str):
            v = v.strip()
            return v if v else None
        return v

    # ─── Импорт одной строки ──────────────────────────────────────────

    def _import_row(self, sheet, row_idx: int, h: dict[str, int | None], *,
                    historical_map: HistoricalMap, informant: Person | None):
        """Импортирует одну строку → Place + (1-2) Toponym."""

        name_evn = self._cell(sheet, row_idx, h.get("name_evn"))
        name_lat = self._cell(sheet, row_idx, h.get("name_lat"))
        translation_ru = self._cell(sheet, row_idx, h.get("translation_ru"))
        translation_en = self._cell(sheet, row_idx, h.get("translation_en"))
        affix = self._cell(sheet, row_idx, h.get("affix"))
        obj_ru = self._cell(sheet, row_idx, h.get("obj_ru"))
        obj_en = self._cell(sheet, row_idx, h.get("obj_en"))
        obj_evn_cyr = self._cell(sheet, row_idx, h.get("obj_evn_cyr"))
        obj_evn_lat = self._cell(sheet, row_idx, h.get("obj_evn_lat"))
        lang_raw = self._cell(sheet, row_idx, h.get("language"))
        comment = self._cell(sheet, row_idx, h.get("comment"))
        official_name = self._cell(sheet, row_idx, h.get("official_name"))
        lat = self._cell(sheet, row_idx, h.get("latitude"))
        lng = self._cell(sheet, row_idx, h.get("longitude"))
        precise = self._cell(sheet, row_idx, h.get("precise"))
        source_raw = self._cell(sheet, row_idx, h.get("source"))

        # Язык: нормализация
        iso = LANGUAGE_NORMALIZATION.get(str(lang_raw)) if lang_raw else "evn"
        language = self.languages.get(iso or "evn", self.languages["evn"])

        # Тип объекта
        feature_type = self._get_or_create_feature_type(
            ru=str(obj_ru) if obj_ru else None,
            en=str(obj_en) if obj_en else None,
            evn_cyr=str(obj_evn_cyr) if obj_evn_cyr else None,
            evn_lat=str(obj_evn_lat) if obj_evn_lat else None,
        )

        # Координаты
        latitude, longitude = self._parse_coords(lat, lng)
        if latitude is None and longitude is None:
            self.stats.rows_no_coords += 1
        is_approximate = (str(precise) == "0") if precise is not None else False

        # Источник координат — как комментарий к Place
        loc_comment_parts = []
        if source_raw:
            normalized_source = COORD_SOURCE_NORMALIZATION.get(str(source_raw), str(source_raw))
            loc_comment_parts.append(f"Источник координат: {normalized_source}")
        if comment:
            loc_comment_parts.append(str(comment))
        location_comment = "; ".join(loc_comment_parts)

        # Создаём Place
        if self.dry_run:
            self.stats.places_created += 1
            return

        place = Place(
            latitude=latitude,
            longitude=longitude,
            feature_type=feature_type or self._fallback_feature_type(),
            historical_map=historical_map,
            is_coordinates_approximate=is_approximate,
            location_comment=location_comment,
        )
        place.save()
        self.stats.places_created += 1

        # Создаём эвенкийский (или иной) Toponym
        toponym = Toponym(
            place=place,
            name=str(name_evn),
            language=language,
            translation_ru=str(translation_ru) if translation_ru else "",
            translation_en=str(translation_en) if translation_en else "",
            linguistic_means=f"Аффикс: {affix}" if affix else "",
            source=self.archival_source,
            historical_map=historical_map,
            informant=informant,
        )
        # Если язык эвенкийский — заполнить латиницу из таблицы (если есть)
        if language.iso == "evn" and name_lat:
            toponym.name_latin = str(name_lat)

        toponym.save()
        self.stats.toponyms_evn_created += 1

        # Официальное русское имя — отдельный Toponym с тем же Place
        if official_name:
            ru_toponym = Toponym(
                place=place,
                name=str(official_name),
                language=self.languages["ru"],
                source=self.archival_source,
                historical_map=historical_map,
                informant=informant,
            )
            ru_toponym.save()
            self.stats.toponyms_ru_created += 1

    # ─── Хелперы ──────────────────────────────────────────────────────

    def _parse_coords(self, lat: Any, lng: Any) -> tuple[Decimal | None, Decimal | None]:
        """Парсит координаты — поддерживает запятую и точку как разделитель."""
        return self._parse_one_coord(lat), self._parse_one_coord(lng)

    def _parse_one_coord(self, val: Any) -> Decimal | None:
        if val is None or val == "":
            return None
        if isinstance(val, (int, float)):
            return Decimal(str(val))
        # Строка с запятой: "60,67925" → "60.67925"
        s = str(val).strip().replace(",", ".")
        try:
            return Decimal(s)
        except (InvalidOperation, ValueError):
            return None

    def _get_or_create_person(self, raw: str | None) -> Person | None:
        """Создаёт информанта.

        Имена информантов в данных Василевич идут в разных форматах
        ("Колесов Федор Николаевич", "Сафронов Федот", "Н.Н. Пудов (Бута)",
        "Марков Петр Алексеевич и Кэптуке Николай"). Парсить их в
        first/patronymic/last_name надёжно невозможно — кладём целое имя
        в last_name, а пометки и латиницу в comment.
        """
        if not raw:
            return None
        raw_str = str(raw).strip()
        if not raw_str or raw_str.lower() in ("unknown", "неизвестно"):
            return None

        # Латинская транслитерация в скобках уходит в комментарий
        translit_match = re.search(r"\(([^)]+)\)", raw_str)
        translit = translit_match.group(1).strip() if translit_match else ""
        clean = re.sub(r"\s*\([^)]+\)\s*", " ", raw_str).strip()
        # Отделяем годы типа ", 1929"
        year_match = re.search(r",\s*(\d{4}.*?)$", clean)
        year_info = year_match.group(1).strip() if year_match else ""
        clean = re.sub(r",\s*\d{4}.*$", "", clean).strip()

        comment_parts = [f"Оригинальная запись: {raw_str}"]
        if translit:
            comment_parts.append(f"Транслитерация: {translit}")
        if year_info:
            comment_parts.append(f"Год: {year_info}")
        comment = "; ".join(comment_parts)

        if self.dry_run:
            return Person(last_name=clean, first_name="", patronymic="", comment=comment)

        person, _ = Person.objects.get_or_create(
            last_name=clean,
            first_name="",
            patronymic="",
            defaults={"comment": comment},
        )
        return person

    def _get_or_create_map(self, map_number: str, archive_ref: str,
                          informant: Person | None) -> HistoricalMap | None:
        """Создаёт или находит HistoricalMap по архивной ссылке."""
        # Используем полную архивную ссылку для уникальности
        # (map 17 встречается 4 раза с разными archive_ref)
        area_name = f"Карта {map_number} ({archive_ref}, Василевич)"
        area_name_en = f"Map {map_number} ({archive_ref}, Vasilevich)"

        if self.dry_run:
            self.stats.maps_created += 1
            return HistoricalMap(area_name_ru=area_name, area_name_en=area_name_en)

        # Ищем по уникальному комментарию с архивной ссылкой
        existing = HistoricalMap.objects.filter(area_name_ru=area_name).first()
        if existing:
            if not self.update_existing:
                self.stats.maps_skipped += 1
                return existing
            self.stats.maps_updated += 1
            # Удалим связанные Place — они будут заново создаваться,
            # каскадно почистятся Toponym
            existing.places.all().delete()
            return existing

        hm = HistoricalMap(
            area_name_ru=area_name,
            area_name_en=area_name_en,
            author=informant,
            collector=self.collector,
            is_archive=True,
            comment_collected_ru=f"Архивная ссылка: {archive_ref}",
            comment_collected_en=f"Archive reference: {archive_ref}",
            place_collected=ARCHIVAL_SOURCE,
        )
        hm.save()
        self.stats.maps_created += 1
        return hm

    def _fallback_feature_type(self) -> FeatureType:
        """FeatureType по умолчанию, если в строке не указан."""
        ft, _ = FeatureType.objects.get_or_create(
            code="unknown",
            defaults={"name_ru": "Неизвестно", "name_en": "Unknown"},
        )
        return ft

    # ─── Лог ───────────────────────────────────────────────────────────

    def log(self, msg: str):
        self.log_lines.append(msg)
        self.stdout.write(msg)

    def _log_header(self, t: str):
        self.log(f"\n## {t}\n")

    def _write_summary(self):
        self._log_header("Сводка")
        s = self.stats
        self.log(f"- Вкладок обработано: **{s.sheets_processed}**")
        self.log(f"- Рукописных карт создано: **{s.maps_created}**, обновлено: {s.maps_updated}, пропущено: {s.maps_skipped}")
        self.log(f"- Мест (Place) создано: **{s.places_created}**")
        self.log(f"  - из них без координат: {s.rows_no_coords}")
        self.log(f"- Топонимов (Toponym) создано:")
        self.log(f"  - эвенкийских/якутских/русских из колонки name: **{s.toponyms_evn_created}**")
        self.log(f"  - русских из Official names: **{s.toponyms_ru_created}**")
        self.log(f"  - всего: **{s.toponyms_evn_created + s.toponyms_ru_created}**")
        self.log(f"- Пустых строк пропущено: {s.rows_skipped_empty}")
        if s.warnings:
            self.log(f"\n**Предупреждений: {len(s.warnings)}**")
            for w in s.warnings[:30]:
                self.log(f"  - {w}")
            if len(s.warnings) > 30:
                self.log(f"  - ... и ещё {len(s.warnings) - 30}")

    def _save_log(self):
        self.log_output.parent.mkdir(parents=True, exist_ok=True)
        self.log_output.write_text("\n".join(self.log_lines), encoding="utf-8")
